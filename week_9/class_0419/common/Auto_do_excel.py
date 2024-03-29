#-*- coding:utf-8 _*-
"""
  @author:小胡
  @file: Auto_do_excel.py
  @time: 2019/04/15
  
  """
import openpyxl

from week_8.class_0413.common import Task_http_request


class Case:
    """
    测试用例类，每个测试用例，实际上就是它的一个实例
    """
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.workbook=openpyxl.load_workbook(file_name)#打开excel
        self.sheet_name=sheet_name
        self.sheet=self.workbook[sheet_name]

    def get_cases(self):
        max_row=self.sheet.max_row#获取最大行数
        cases=[]#存放所有的测试用例
        for r in range(2,max_row+1):
            # case={}
            # case['case_id']=self.sheet.cell(row=r,column=1)
            # case['title']=self.sheet.cell(row=r,column=2)
            case=Case()
            case.case_id=self.sheet.cell(row=r,column=1).value
            case.title=self.sheet.cell(row=r,column=2).value
            case.url=self.sheet.cell(row=r,column=3).value
            case.data=self.sheet.cell(row=r,column=4).value
            case.method=self.sheet.cell(row=r,column=5).value
            case.expected=self.sheet.cell(row=r,column=6).value
            cases.append(case)
            self.workbook.close()
        return cases#返回cases列表



    def write_result(self,row,actual,result):
        sheet=self.workbook[self.sheet_name]
        sheet.cell(row,7).value=actual
        sheet.cell(row,8).value=result
        # self.workbook.save(filename=self.file_name)
        self.workbook.save(filename=self.file_name)
        self.workbook.close()




if __name__ == '__main__':
    from week_8.class_0413.common import contants
    do_excel=DoExcel(contants.case_file,sheet_name='recharge')
    cases=do_excel.get_cases()
    http_request= Task_http_request.HttpRequest()

    for case in cases:
        # print(case.case_id)
        # print(case.method)
        # print(case.data)
        print(case.__dict__)
        resp=http_request.request(case.method,
                             case.url,
                             case.data)
        print(resp.status_code)
        print(resp.text)  # 响应文本
        resp_dict = resp.json()  # 返回字典
        print(resp_dict)
        actual=resp.text
        if case.expected==actual:#判断期望结果是否与实际结果一致
            do_excel.write_result(case.case_id+1,actual,'PASS')
        else:
            do_excel.write_result(case.case_id+1,actual,'FAIL')



