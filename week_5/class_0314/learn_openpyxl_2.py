# -*- coding: utf-8 -*-
# @Time    : 2019/3/14 21:41
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : learn_openpyxl_2.py

#读诗  静夜思
from openpyxl import load_workbook

wb=load_workbook('py15.xlsx')

sheet=wb['Sheet1']

# print(sheet.max_row)#row 获取最大行
# print(sheet.max_column)#column 获取最大列
for i in range(1,sheet.max_row+1):#控制行
    for j in range(1,sheet.max_column+1):#1 2 3 4  为什么要加1？#控制列

        if sheet.cell(i,j).value:#???
            print(sheet.cell(i,j).value)
        # print(sheet.cell(2,2).value)
        # print(sheet.cell(2,3).value)
        # print(sheet.cell(2,4).value)

#单元格为空的时候  读取到的是None


#安排一个作业
#写一个类  类里面有2个方法  1）读数据  2）写数据
#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，所有子列表数据放到一个大列表里面,要求把读取到的数据返回
#2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值
#温馨提示：记得关闭和保存Excel
