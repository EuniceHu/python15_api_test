#-*- coding:utf-8 _*-
"""
  @author:小胡
  @file: class_openpyxl_3.py
  @time: 2019/03/18
  
  """
#安排一个作业
#写一个类  类里面有2个方法  1）读数据  2）写数据
#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，
# 所有子列表数据放到一个大列表里面,要求把读取到的数据返回
#2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值
#温馨提示：记得关闭和保存Excel

# 分析：类里面有2个方法  读写，读的数据里面有2个列表，使用的列表之间的相互追加
# 首先要是编写类的话，就把一切能参数化的，都要参数化，这样更能更好的调用和复用
# 比如文件的名称，sheet表单名  写数据的行，列，值参数化
#导入读写excel的模块
from openpyxl import load_workbook

#定义一个类

class ReadWrite:
    '''这是一个读写类'''
    #初始化参数
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        test_data=[]
        for i in range(1,sheet.max_row+1):
            test_data_1 = []
            for j in range(1,sheet.max_column+1):
                item=sheet.cell(i,j).value
                test_data_1.append(item)
            test_data.append(test_data_1)
        return test_data

    def write(self,row,column,value):
        #打开工作簿
        wb=load_workbook(self.file_name)
        #定位表单
        sheet=wb[self.sheet_name]
        sheet.cell(row,column).value=value
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    t=ReadWrite('py_15_1.xlsx','lhh')
    res = t.read()
    print(res)
    t.write(row=7,column=2,value='火烈鸟')