#-*- coding:utf-8 _*-
"""
  @author:小胡
  @file: class_openpyxy.py
  @time: 2019/03/14
  
  """
#创建excel文件的模块workbook
#读写excel文件的模块 Load_work
# 创建一个Excel：手动  代码创建
# 新建一个Excel
# from openpyxl import workbook#导入workbook模块
# wb=workbook.Workbook()
# wb.create_sheet('lhh',index=0)#创建表单的方法，创建一个自己命名的表单
# wb.save('py_15_1.xlsx')#另存为 保存工作簿



#开始操作Excel 读写
#读写Excel 三步骤
#第一：先打开Excel
from openpyxl import load_workbook
wb=load_workbook('py_15_1.xlsx')#打开工作簿

#第二：定位到表单
sheet=wb['lhh']

#第三：定位到单元格 获取内容  根据行列坐标来定位单元格再获取值
#行列必须是数字 对应关系：A-1 B-2 C-3 D-4 E-5

# A1=sheet.cell(1,1).value
# print('A1的值为：{}，A1的类型为：{}'.format(A1,type(A1)))
#
# A3=sheet.cell(3,1).value
# print('A3的值为：{}，A3的类型为：{}'.format(A3,type(A3)))
#
# B3=sheet.cell(3,2).value
# print('B3的值为：{}，B3的类型为：{}'.format(B3,type(B3)))
#
#
# C3=sheet.cell(3,3).value
# print('C3的值为：{}，C3的类型为：{}'.format(C3,type(C3)))
#

C4=sheet.cell(4,3).value
print('C3的值为：{}，C3的类型为：{}'.format(C4,type(C4)))

C4=eval(sheet.cell(4,3).value)
print('C3的值为：{}，C3的类型为：{}'.format(C4,type(C4)))

# eval()可以把数据转换为Python原本可以识别的数据类型
# 读写excel中的数据，可以得出一个结论
# 数字还是数字，其他数据类型全是字符串类型


# 写入值  保存
#写入新的值
sheet.cell(6,1).value='今天礼拜一'

#替换值
sheet.cell(3,3,'hhh')


# 保存工作簿
wb.save('py_15_1.xlsx')

#关闭工作簿

wb.close()