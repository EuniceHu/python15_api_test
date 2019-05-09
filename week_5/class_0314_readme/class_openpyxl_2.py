#-*- coding:utf-8 _*-
"""
  @author:小胡
  @file: class_openpyxl_2.py
  @time: 2019/03/18
  
  """


#读Excel

#第一：打开工作簿

from openpyxl import  load_workbook

wb=load_workbook('py_15_1.xlsx')

#第二：定位表单

sheet=wb['lhh']

#第三：定位单元格

A1=sheet.cell(1,1).value
print(A1)

print(sheet.max_row)#获取最大行
print(sheet.max_column)#获取最大列


for i in range(1,sheet.max_row+1):#1,2,3,4,5,6
    for j in range(1,sheet.max_column+1):#1,2,3,4
        if sheet.cell(i,j).value:#if后面的条件如果为真，就继续执行
            print(sheet.cell(i,j).value)

#单元格为空的时候  读取到的是None


