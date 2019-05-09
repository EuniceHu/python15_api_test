# encoding=utf_8
# @Author  ： 豆子
from openpyxl import load_workbook

from douzi_0323.commons import filepath


class ExcelParse(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(filename)

    def get_cases(self, sheetname):
        sheet = self.wb[sheetname]
        test_cases = []
        for row in range(2, sheet.max_row+1):
            row_case = {}
            row_case['case_id'] = sheet.cell(row, 1).value
            row_case['title'] = sheet.cell(row, 2).value
            row_case['url'] = sheet.cell(row, 3).value
            row_case['method'] = sheet.cell(row, 4).value
            row_case['param'] = sheet.cell(row, 5).value
            row_case['excepted'] = sheet.cell(row, 6).value
            test_cases.append(row_case)
        return test_cases

    def back_write_by_excel(self, sheetname, case_id, actual, result):
        sheet = self.wb[sheetname]
        sheet.cell(row=case_id+1, column=7, value=actual)
        sheet.cell(row=case_id+1, column=8, value=result)
        self.wb.save(self.filename)
        self.wb.close()


if __name__ == '__main__':
    ep = ExcelParse(filepath.data_dir)
    cases = ep.get_cases('test_data')
    for case in cases:
        print(case)






