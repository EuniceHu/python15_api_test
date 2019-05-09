# encoding: utf-8 
"""
@author: 凹凸曼
@number: 1526
@file: homework_0323.py
@time: 2019/3/24 20:43
"""
import unittest,HTMLTestRunner
from week_5.homework_0312 import UserLogin
from ddt import ddt,data,unpack
from openpyxl import load_workbook

#方法一
@ddt
class TestLogin(unittest.TestCase):

    case_data = [
        {
            "url": "http://47.107.168.87:8080/futureloan/mvc/api/member/login",
            "params": {'mobilephone': '18688773467', 'pwd': '123456'},
            "method": "get"
        },
        {
            "url": "http://47.107.168.87:8080/futureloan/mvc/api/member/login",
            "params": {'mobilephone': '18688773467', 'pwd': '123456'},
            "method": "post"
        },
        {
            "url": "http://47.107.168.87:8080/futureloan/mvc/api/member/login",
            "params": {'mobilephone': '18688773467', 'pwd': '12345678'},
            "method": "post"
        },
        {
            "url": "http://47.107.168.87:8080/futureloan/mvc/api/member/login",
            "params": {'mobilephone': '13000001111', 'pwd': '123456'},
            "method": "post"
        }
    ]
    @data(*case_data)
    def test_user_login_001(self,user_info):
        expected = '登录成功'
        res = UserLogin(user_info["url"], user_info["params"],user_info["method"]).HttpRequest()['msg']
        self.assertEqual(expected, res)

#方法二：从Excel表获取测试数据，并执行测试
@ddt
class TestLogin(unittest.TestCase):
    wb = load_workbook('LoginData.xlsx')  # 打开工作簿
    sheet = wb['Sheet1']  # 定位表单
    dic = []  # 定义空字典
    for i in range(2, sheet.max_row + 1):  # 2个嵌套for循环,读取多行多列，从第二行开始读取
        d = []  # 定义空字典
        for j in range(1, sheet.max_column + 1):  # sheet.max_row获取最大行，sheet.max_column获取最大列
            if sheet.cell(i, j).value:  # 判断非空数据才执行
                resp = sheet.cell(i, j).value
                d.append(resp)
        dic.append(d)
    @data(*dic)   # 将列表的每一个元素作为参数参数传入到dada里
    @unpack       #将data里的数据再次进行解包
    def test_user_login_001(self,a,b,user_info):
        param=eval(user_info)
        expected = '登录成功'           #期望值

        # 调用UserLogin类里的HttpRequest方法并读取响应报文里“msg”的value值
        res = UserLogin(b, param, a).HttpRequest()['msg']
        self.assertEqual(expected, res)                     #判断期望值与实际值是否一致
