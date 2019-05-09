from openpyxl import load_workbook


class PlayExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read_data(self):

        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]


        test_data=[]

        for i in range(1,sheet.max_row+1):
            sub_data={}
            sub_data['http_method'] = sheet.cell(i, 1).value
            sub_data['url'] = sheet.cell(i, 2).value
            sub_data['param'] = sheet.cell(i, 3).value
            sub_data['Expected'] = sheet.cell(i, 4).value
            test_data.append(sub_data)


        wb.close()
        return test_data


if __name__ == '__main__':
   print(PlayExcel('http_request_data.xlsx','http_request_data').read_data())




