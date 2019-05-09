import unittest
from work_08 import Request
from flask import json
from openpyxl import load_workbook

class TestGetRequest(unittest.TestCase):

    wb = load_workbook('testdata_0321.xlsx')
    sheet =wb['请求数据']

    def test_phone_pwd_correct(self):#验证get请求，账号密码正确时，响应的msg的正确性
        expected = self.sheet.cell(2,5).value
        res = Request().http_request(self.sheet.cell(2,1).value,eval(self.sheet.cell(2,2).value),self.sheet.cell(2,3).value)
        actual = json.loads(res[3])#用json.loads()方法将字符串转换为字典，并且这个方法会将字典中的null转换为none
        self.assertEqual(expected,actual['msg'])

    def test_pwd_wrong(self):#验证get请求，账号正确密码错误时，响应的msg的正确性
        expected = self.sheet.cell(3,5).value
        res = Request().http_request(self.sheet.cell(3,1).value,eval(self.sheet.cell(3,2).value),self.sheet.cell(3,3).value)
        actual = json.loads(res[3])
        self.assertEqual(expected,actual['msg'])

    wb.close()

class TestPostRequest(unittest.TestCase):

    wb = load_workbook('testdata_0321.xlsx')
    sheet = wb['请求数据']

    def test_phone_pwd_correct(self):#验证post请求，账号密码正确时，响应的status的正确性
        expected = self.sheet.cell(4,4).value
        res = Request().http_request(self.sheet.cell(4,1).value,eval(self.sheet.cell(4,2).value),self.sheet.cell(4,3).value)
        actual = json.loads(res[3])
        self.assertEqual(expected,actual['status'])

    def test_pwd_wrong(self):#验证post请求，账号正确密码错误时，响应的status的正确性
        expected = self.sheet.cell(5,4).value
        res = Request().http_request(self.sheet.cell(5,1).value,eval(self.sheet.cell(5,2).value),self.sheet.cell(5,3).value)
        actual = json.loads(res[3])
        self.assertEqual(expected,actual['status'])

    wb.close()