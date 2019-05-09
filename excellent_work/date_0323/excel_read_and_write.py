#coding:utf-8
#@time : 2019/3/24 17:43
# @Author : apple
#@file : excel_read_and_write.py

from openpyxl import load_workbook
class ExcelReadAndWrite:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel(self,row_end_id,column_end_id,row_start_id=1,column_start_id=1):
        '''

        :file_name1: 文件簿名称
        :sheet_name1: 表单名
        '''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        all_data = []
        for i in range(row_start_id,row_end_id+1):
            sub_list = []
            for j in range(column_start_id,column_end_id+1):
                item = sheet.cell(i,j).value
                # print(item)
                sub_list.append(item)
            all_data.append(sub_list)

        wb.close()#关闭文件
        return all_data#返回测试数据

    def write_excel(self, row_id, column_id, content):
        wb = load_workbook(self.file_name)#打开工作簿
        sheet = wb[self.sheet_name]#定位到表单
        sheet.cell(row_id, column_id, content)  # 指定位置赋值写入

        wb.save(self.file_name)
        wb.close()

if __name__ == '__main__':
    file_name='test_data.xlsx'
    #write_data = ExcelReadAndWrite(file_name,'Sheet2').write_excel(1,2,'床前明月光')
    read_data = ExcelReadAndWrite(file_name,'Sheet1').read_excel(5,4,row_start_id=2)
    print("read后的返回值是：",read_data)


