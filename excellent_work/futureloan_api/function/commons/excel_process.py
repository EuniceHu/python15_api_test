# -*- coding: utf-8 -*-
# @Author  : hubing
# @mail    : cocofei-f@163.com
# @Date    : 2019-04-15 16:21
from openpyxl import load_workbook

from function.commons.base import Base
from function.commons.logger import Logger


class ExcelProcess():

    def __init__(self, filename):
        logger = Logger("file")
        try:
            self.wb = load_workbook(filename)
            self.filename = filename
        except BaseException as e:
            print('打开excel文件失败：',e)
            logger.getLoger().error('打开excel工作表失败!',e)

    def save(self, filename=None):
        '''
        保存excel文件，默认保持原始文件，可选另保存
        :param filename:
        :return:
        '''
        try:
            if filename:
                self.wb.save(filename)
            else:
                self.wb.save(self.filename)

        except BaseException as e:
            print('保存excel文件失败！',e)


    def close(self):
        '''关闭excel文件'''
        self.wb.close()

    def read_excel_sheet(self, sheetName=None):
        """
        默认读取excel第一张工作表数据，也可指定sheet读取数据
        :param sheet:
        :return:
        """
        if sheetName:
            sheet = self.wb[sheetName]
        else:
            sheet = self.wb.worksheets[0]

        data = sheet.rows

        list_all = []
        for i in data:
            list_row = []
            for j in i:
                list_row.append(j.value)
            list_all.append(list_row)
        self.close()
        return list_all

    def read_excel_sheet_row(self, sheetName=None, rows=0):
        '''
        默认获取
        :param sheetName:默认读取excel第一张工作表数据，也可指定sheet读取数据
        :param row:默认读取execl首行，可输入行数读取制定起始行数据
        :return:二维list
        '''
        if sheetName:  # 选择传入的sheet工资表
            sheet = self.wb[sheetName]
        else:  # 如果sheet没有值，则选择wb的第一个工作表
            sheet = self.wb.worksheets[0]

        max_row = sheet.max_row
        max_column = sheet.max_column
        # 定义list_all保存工作表每行数据
        list_all = []
        for i in range(1 + rows, max_row + 1):
            rowList = []
            for j in range(1, max_column):
                rowList.append(sheet.cell(i, j).value)
            list_all.append(rowList)
        self.close()
        return list_all

    def read_excel_all(self):
        '''
        读取excel文件中所有的工作表数据
        :return:三维list列表
        '''
        data = []
        for i in self.wb:
            sheet_list = []
            for j in i:
                row_list = []
                for k in j:
                    row_list.append(k.value)
                sheet_list.append(row_list)
            data.append(sheet_list)
        self.close()
        return data

    def write_excel_cell(self, cell, value, sheet=None):
        '''
        填写数据到单元格
        :param cell:    单元格，例如[A1]、[B4]
        :param value:   数据
        :param sheet:   工资表，默认第一个工作表，可指定sheet
        :param file:    默认保存原始文件，可输入文件名另保存
        :return:
        '''
        if sheet:
            sheet = self.wb[sheet]
        else:
            sheet = self.wb.worksheets[0]

        sheet[cell] = value

    def write_excel_rowColumn(self,row,column,value,sheet=None):
        '''
        填写数据到指定单元格
        :param row: 工作表行数
        :param column:  工资表列数
        :param value:   数据
        :param sheet:   工资表，默认第一个工作表，可指定sheet
        :param file:    默认保存原始文件，可输入文件名另保存
        :return:
        '''
        if sheet:
            sheet = self.wb[sheet]
        else:
            sheet = self.wb.worksheets[0]
        sheet.cell(row, column, value)


if __name__ == '__main__':
    excel  = ExcelProcess(Base.get_cur_dir(__file__+'/../../../data/futureloan_cases.xlsx'))
    excel.wb