#-*- coding:utf-8 _*-
"""
  @author:小胡
  @file: Task_openpyxl_testcase.py
  @time: 2019/04/02
  
  """
import unittest
from openpyxl import load_workbook#导入操作Excel的模块
from ddt import ddt,data,unpack#导入操作ddt的模块
from Task.class_http_requests.class_requests_one import UserLogin
#导入http_requests的类


#编写一个test类
@ddt#用ddt做修饰
class TestHttpRequeste(unittest.TestCase):
    #打开Excel
    wb=load_workbook('Homework_ddt.xlsx')
    sheet=wb.worksheets[0]#定位表单
    list=[]#定义一个空列表
    for i in range(2,sheet.max_row+1):
        list_1=[]
        for j in range(1,sheet.max_column+1):
            item=sheet.cell(i,j).value#读取Excel中的数据
            list_1.append(item)
        list.append(list_1)
    wb.close()
    @data(*list)#将列表中的数据传到data里面
    @unpack#进行分割
    def test_001(self,a,url,login_info,excepted):
        params=eval(login_info)
        excepted=excepted
        res=UserLogin(url,params,a).http_request()['msg']
        self.assertEqual(excepted,res)

