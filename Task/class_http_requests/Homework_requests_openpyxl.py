#-*- coding:utf-8 _*-
"""
  @author:小胡
  @file: Homework_requests_openxy.py
  @time: 2019/03/25
  
  """
import unittest
from ddt import ddt,data,unpack
from Task.class_http_requests.class_requests import UserLogin
from openpyxl import load_workbook

#把测试数据存放在excel文件中，写测试用例的时候，要把Excel中的数据给拿出来
#先导入模块，读取Excel，定位到表单
@ddt
class TestHttpRequests(unittest.TestCase):

    wb=load_workbook('Homework_ddt.xlsx')#打开工作簿
    sheet=wb['测试数据']#定位表单
    # sheet=wb.worksheets[0]
    list=[]#定义空列表
    # for i in range(2,sheet.max_row+1):
    for i in range(2,sheet.max_row+1):#读取多行多列，从第二行开始读取
        list_1=[]
        for j in range(1, sheet.max_column + 1):
            item=sheet.cell(i,j).value
            list_1.append(item)
        list.append(list_1)
    wb.close()
    @data(*list)#将列表里面的元素传到data里
    @unpack#进行分割
    def test_001(self,a,url,login_info,expected):
        params=eval(login_info)
        expected=expected
        #调用HttpRequests的方法
        res=UserLogin(url,params,a).HttpRequest()['msg']
        self.assertEqual(expected,res)
#['msg']是响应报文的值，首先报文里面数据类型是字典，那么要判断期望值与实际值是否一致 dict['msg']
#msg相当于key值，dict['msg']取Value，Value就是实际值
#只需要判断期望值与实际值是否相等

