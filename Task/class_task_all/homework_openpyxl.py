#-*- coding:utf-8 _*-
"""
  @author:小胡
  @file: homework_openpyxl.py
  @time: 2019/03/16
  
  """
#安排一个作业 #写一个类 类里面有2个方法 1）读数据 2）写数据

#1）读数据可以读取整个Excel里面所有的数据，每一行数据都放到一个子列表里面，
# 所有子列表数据放到一个大列表里面,要求把读取到的数据返回
# #2）写数据可以在Excel里面指定的单元格里面写入指定的值，不需要返回值
#温馨提示：记得关闭和保存Excel


from openpyxl import load_workbook

class DoExcel:

    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name




    def read(self):
        #打开工作簿，定位到表单
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        #定义一个列表
        test_data=[]
        for i in range(1,sheet.max_row+1):
              test_data_1=[]
              for j in range(1,sheet.max_column+1):
                  item = sheet.cell(i,j).value
                  test_data_1.append(item)
              test_data.append(test_data_1)
        return test_data



    # def read(self):
    #     #打开工作簿，定位到表单
    #     wb=load_workbook(self.file_path)
    #     sheet=wb[self.sheet_name]
    #     #定义一个列表
    #     test_data=[]
    #     for i in range(2,sheet.max_row+1):
    #       test_data_1={}
    #       test_data_1['id'] = sheet.cell(i,1).value
    #       test_data_1['name'] = sheet.cell(i, 2).value
    #       test_data_1['title'] = sheet.cell(i, 3).value
    #       test_data_1['hoppy'] = sheet.cell(i, 4).value
    #       test_data.append(test_data_1)
    #     return test_data


    #
    # def write(self,row,column,value):
    #     '''
    #     :param row: 写入的行数
    #     :param column: 写入的列数
    #     :param value: 写入的值
    #     :return:
    #     '''
    #     wb=load_workbook(self.file_path)
    #     sheet=wb[self.sheet_name]
    #     sheet.cell(row,column).value=value
    #     wb.save(self.file_path)
    #     wb.close()

if __name__ == '__main__':
    t=DoExcel('Homework.xlsx','test').read()
    print(t)
    # t1=DoExcel('Homework.xlsx','test')
    # t1.write(row=6,column=1,value='5')


